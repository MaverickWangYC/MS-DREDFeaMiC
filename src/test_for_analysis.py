from sklearn.neural_network import MLPClassifier
from sklearn.model_selection import StratifiedShuffleSplit
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, roc_auc_score
from h5_reader import read_h5
import os
from read_OC import read_lowResolution_8_7_02
from read_MI import read_MI_batch, read_MI_batch_after_Ka
from read_CHD import read_CHD_batch
from read_mouse import read_mouse
from read_CC import read_CC_batch
from read_CRLM import read_CRLM
from read_tomato import read_tomato
from read_CHD_MSFC import read_CHD_MSFC
from read_KidneyCancer import read_KidneyCancer
from read_KidneyDisease import read_KidneyDisease
from read_Pneumonia import read_Pneumonia

import math
import pandas as pd
import torch
from torch import nn
from torch.utils import data
import shap
import numpy as np
from model import (E_D_Mamba)
from torch.optim.lr_scheduler import CosineAnnealingLR,CosineAnnealingWarmRestarts,StepLR
import torch.nn.functional as F
import random
from sklearn.model_selection import train_test_split
# from d2l import torch as d2l
from torch.nn import functional as F
import pywt
from sklearn.metrics import classification_report
import matplotlib.pyplot as plt
from torch.autograd import Variable
import argparse
from sklearn.preprocessing import MinMaxScaler
import openpyxl
# from DBLoss import BalancedCrossEntropyLoss
# device = "cpu"
device = 'cuda' if torch.cuda.is_available() else 'cpu'
def seed_everything(seed=11):
    # random.seed(seed)
    # np.random.seed(seed)
    # torch.manual_seed(seed)
    # torch.cuda.manual_seed(seed)
    # torch.cuda.manual_seed_all(seed)
    # torch.backends.cudnn.deterministic = True
    # torch.backends.cudnn.benchmark = False
    # 固定随机种子等操作
    seed_n = seed
    print('seed is ' + str(seed_n))
    g = torch.Generator()
    g.manual_seed(seed_n)
    random.seed(seed_n)
    np.random.seed(seed_n)
    torch.manual_seed(seed_n)
    torch.cuda.manual_seed(seed_n)
    torch.cuda.manual_seed_all(seed_n)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False
    torch.backends.cudnn.enabled = False
    torch.use_deterministic_algorithms(True)
    os.environ['CUBLAS_WORKSPACE_CONFIG'] = ':16:8'
    os.environ['PYTHONHASHSEED'] = str(seed_n)  # 为了禁止hash随机化，使得实验可复现。

def get_K_fold_data_LungCancer1(K, i):
    file_dir = '../data/LungCancer1'
    batch_1, labels_1, batch_2, labels_2 = read_lungcancer1(file_dir)

    # batch_1 = batch_1 / np.max(batch_1, axis=1)[:, None]
    # batch_2 = batch_2 / np.max(batch_2, axis=1)[:, None]
    unique_chars, integer_labels_1 = np.unique(labels_1, return_inverse=True)
    unique_chars, integer_labels_2 = np.unique(labels_2, return_inverse=True)


    train_data = batch_1
    train_label = integer_labels_1
    test_data = batch_2
    test_label_true = integer_labels_2

    train_p = sum(train_label)
    train_n = len(train_label) - train_p

    train_data = torch.tensor(
        train_data, dtype=torch.float32
    ).to(device)

    train_label = torch.tensor(
        train_label
    ).to(device)

    test_data = torch.tensor(
        test_data, dtype=torch.float32
    ).to(device)

    test_label_true = torch.tensor(
        test_label_true
    ).to(device)

    all_len = train_n + train_p
    class_weights = torch.tensor(
        [1.0 - float(train_n / all_len), 1.0 - float(train_p / all_len)])

    print(train_data.shape)
    return train_data, train_label, test_data, test_label_true, class_weights.to(device)

def get_K_fold_data_LungCancer2(K, i):
    file_dir = '../data/LungCancer2'
    batch_1, labels_1, batch_2, labels_2 = read_lungcancer2(file_dir)

    # batch_1 = batch_1 / np.max(batch_1, axis=1)[:, None]
    # batch_2 = batch_2 / np.max(batch_2, axis=1)[:, None]
    unique_chars, integer_labels_1 = np.unique(labels_1, return_inverse=True)
    unique_chars, integer_labels_2 = np.unique(labels_2, return_inverse=True)


    train_data = batch_1
    train_label = integer_labels_1
    test_data = batch_2
    test_label_true = integer_labels_2

    train_p = sum(train_label)
    train_n = len(train_label) - train_p

    train_data = torch.tensor(
        train_data, dtype=torch.float32
    ).to(device)

    train_label = torch.tensor(
        train_label
    ).to(device)

    test_data = torch.tensor(
        test_data, dtype=torch.float32
    ).to(device)

    test_label_true = torch.tensor(
        test_label_true
    ).to(device)

    all_len = train_n + train_p
    class_weights = torch.tensor(
        [1.0 - float(train_n / all_len), 1.0 - float(train_p / all_len)])

    print(train_data.shape)
    return train_data, train_label, test_data, test_label_true, class_weights.to(device)

def get_K_fold_data_LungCancer6(K, i):
    file_dir = '../data/LungCancer6'
    batch_1, labels_1, batch_2, labels_2 = read_lungcancer6(file_dir)

    # batch_1 = batch_1 / np.max(batch_1, axis=1)[:, None]
    # batch_2 = batch_2 / np.max(batch_2, axis=1)[:, None]
    unique_chars, integer_labels_1 = np.unique(labels_1, return_inverse=True)
    unique_chars, integer_labels_2 = np.unique(labels_2, return_inverse=True)


    train_data = batch_1
    train_label = integer_labels_1
    test_data = batch_2
    test_label_true = integer_labels_2

    train_p = sum(train_label)
    train_n = len(train_label) - train_p

    train_data = torch.tensor(
        train_data, dtype=torch.float32
    ).to(device)

    train_label = torch.tensor(
        train_label
    ).to(device)

    test_data = torch.tensor(
        test_data, dtype=torch.float32
    ).to(device)

    test_label_true = torch.tensor(
        test_label_true
    ).to(device)

    all_len = train_n + train_p
    class_weights = torch.tensor(
        [1.0 - float(train_n / all_len), 1.0 - float(train_p / all_len)])

    print(train_data.shape)
    return train_data, train_label, test_data, test_label_true, class_weights.to(device)

def get_K_fold_data_MultiDisease(K, i):
    file_dir = '../data/MultiDisease'
    batch_1, labels_1, batch_2, labels_2 = read_multidisease(file_dir)

    # batch_1 = batch_1 / np.max(batch_1, axis=1)[:, None]
    # batch_2 = batch_2 / np.max(batch_2, axis=1)[:, None]
    unique_chars, integer_labels_1 = np.unique(labels_1, return_inverse=True)
    unique_chars, integer_labels_2 = np.unique(labels_2, return_inverse=True)


    train_data = batch_1
    train_label = integer_labels_1
    test_data = batch_2
    test_label_true = integer_labels_2

    train_p = sum(train_label)
    train_n = len(train_label) - train_p

    train_data = torch.tensor(
        train_data, dtype=torch.float32
    ).to(device)

    train_label = torch.tensor(
        train_label
    ).to(device)

    test_data = torch.tensor(
        test_data, dtype=torch.float32
    ).to(device)

    test_label_true = torch.tensor(
        test_label_true
    ).to(device)

    all_len = train_n + train_p
    class_weights = torch.tensor(
        [1.0 - float(train_n / all_len), 1.0 - float(train_p / all_len)])

    print(train_data.shape)
    return train_data, train_label, test_data, test_label_true, class_weights.to(device)

def get_K_fold_data_LungCancer1_feature_analysis(K, i, feature_id):
    file_dir = '../data/LungCancer1'
    batch_1, labels_1, batch_2, labels_2 = read_lungcancer1(file_dir)

    # batch_1 = batch_1 / np.max(batch_1, axis=1)[:, None]
    # batch_2 = batch_2 / np.max(batch_2, axis=1)[:, None]
    unique_chars, integer_labels_1 = np.unique(labels_1, return_inverse=True)
    unique_chars, integer_labels_2 = np.unique(labels_2, return_inverse=True)

    train_data = batch_1
    train_label = integer_labels_1
    test_data = batch_2
    test_label_true = integer_labels_2

    train_p = sum(train_label)
    train_n = len(train_label) - train_p
    np.random.seed(2025)
    np.random.shuffle(train_data[:, feature_id])
    # print(train_data[:, feature_id])
    train_data = torch.tensor(
        train_data, dtype=torch.float32
    ).to(device)

    train_label = torch.tensor(
        train_label
    ).to(device)

    test_data = torch.tensor(
        test_data, dtype=torch.float32
    ).to(device)

    test_label_true = torch.tensor(
        test_label_true
    ).to(device)

    all_len = train_n + train_p
    class_weights = torch.tensor(
        [1.0 - float(train_n / all_len), 1.0 - float(train_p / all_len)])

    print(train_data.shape)
    return train_data, train_label, test_data, test_label_true, class_weights.to(device)

def get_K_fold_data_LungCancer2_feature_analysis(K, i, feature_id):
    file_dir = '../data/LungCancer2'
    batch_1, labels_1, batch_2, labels_2 = read_lungcancer2(file_dir)

    # batch_1 = batch_1 / np.max(batch_1, axis=1)[:, None]
    # batch_2 = batch_2 / np.max(batch_2, axis=1)[:, None]
    unique_chars, integer_labels_1 = np.unique(labels_1, return_inverse=True)
    unique_chars, integer_labels_2 = np.unique(labels_2, return_inverse=True)

    train_data = batch_1
    train_label = integer_labels_1
    test_data = batch_2
    test_label_true = integer_labels_2

    train_p = sum(train_label)
    train_n = len(train_label) - train_p
    np.random.seed(2025)
    np.random.shuffle(train_data[:, feature_id])
    # print(train_data[:, feature_id])
    train_data = torch.tensor(
        train_data, dtype=torch.float32
    ).to(device)

    train_label = torch.tensor(
        train_label
    ).to(device)

    test_data = torch.tensor(
        test_data, dtype=torch.float32
    ).to(device)

    test_label_true = torch.tensor(
        test_label_true
    ).to(device)

    all_len = train_n + train_p
    class_weights = torch.tensor(
        [1.0 - float(train_n / all_len), 1.0 - float(train_p / all_len)])

    print(train_data.shape)
    return train_data, train_label, test_data, test_label_true, class_weights.to(device)

def get_K_fold_data_MultiDisease_feature_analysis(K, i, feature_id):
    file_dir = '../data/MultiDisease'
    batch_1, labels_1, batch_2, labels_2 = read_multidisease(file_dir)

    # batch_1 = batch_1 / np.max(batch_1, axis=1)[:, None]
    # batch_2 = batch_2 / np.max(batch_2, axis=1)[:, None]
    unique_chars, integer_labels_1 = np.unique(labels_1, return_inverse=True)
    unique_chars, integer_labels_2 = np.unique(labels_2, return_inverse=True)

    train_data = batch_1
    train_label = integer_labels_1
    test_data = batch_2
    test_label_true = integer_labels_2

    train_p = sum(train_label)
    train_n = len(train_label) - train_p
    np.random.seed(2025)
    np.random.shuffle(train_data[:, feature_id])
    # print(train_data[:, feature_id])
    train_data = torch.tensor(
        train_data, dtype=torch.float32
    ).to(device)

    train_label = torch.tensor(
        train_label
    ).to(device)

    test_data = torch.tensor(
        test_data, dtype=torch.float32
    ).to(device)

    test_label_true = torch.tensor(
        test_label_true
    ).to(device)

    all_len = train_n + train_p
    class_weights = torch.tensor(
        [1.0 - float(train_n / all_len), 1.0 - float(train_p / all_len)])

    print(train_data.shape)
    return train_data, train_label, test_data, test_label_true, class_weights.to(device)




def get_net(in_features, state_size, median_size, output_size, model_name):
    # net = nn.Sequential(nn.Linear(in_features, 3)).to(device)
    # net = nn.Sequential(MambaSkipLinear2(seq_len=1, d_model=in_features, state_size=64, output_size=3)).to(device)
    # net = nn.Sequential(MambaSkipLinear(seq_len=1, d_model=in_features, median_size=median_size, state_size=state_size, output_size=output_size)).to(device)
    # net = nn.Sequential(Mamba(seq_len=1, d_model=in_features, median_size = median_size, state_size=state_size, output_size=3)).to(device)
    # net = TF(in_features=in_features, drop=0.).to(device)
    if model_name == "LSTM":
        net = LSTM(input_size=in_features, hidden_size=median_size, output_size=output_size).to(device)
    if model_name == "RNN":
        net = RNN(input_size=in_features, hidden_size=median_size, output_size=output_size).to(device)
    # net = nn.Sequential(baselinemlp(in_features, 128)).to(device)
    # net = nn.Sequential(TransMamba(seq_len=1, d_model=in_features, median_size=median_size, state_size=state_size, output_size=output_size)).to(device)
    # net = nn.Sequential(Dense_MISM_Trans_MoE(seq_len=1, d_model=in_features, median_size=median_size, state_size=state_size, output_size=output_size)).to(device)
    # net = nn.Sequential(DenseNet(init_features=in_features,classes=output_size))
    if model_name == "E_D_IS_Mamba":
        net = nn.Sequential(E_D_IS_Mamba(seq_len=1, d_model=in_features, median_size=median_size, state_size=state_size, output_size=output_size)).to(device)
    # net = nn.Sequential(Dense_MISM_Trans_MoE(seq_len=1, d_model=in_features, median_size=median_size, state_size=state_size,output_size=output_size)).to(device)
    # net = nn.Sequential(DenseNet(init_features = in_features, classes = output_size)).to(device)
    if model_name == "E_D_Mamba_E_D":
        net = E_D_Mamba_E_D(seq_len=1, d_model=in_features, median_size=median_size, state_size=state_size,
                                      output_size=output_size).to(device)
    if model_name == "E_D_Mamba":
        net = E_D_Mamba(seq_len=1, d_model=in_features, median_size=median_size, state_size=state_size,
                                      output_size=output_size).to(device)

    if model_name == "E_D_Mamba_NoSample":
        net = E_D_Mamba_NoSample(seq_len=1, d_model=in_features, median_size=median_size, state_size=state_size,
                                      output_size=output_size).to(device)
    if model_name == "E_D_Transformer":
        net = E_D_Transformer(seq_len=1, d_model=in_features, median_size=median_size, state_size=state_size,
                                      output_size=output_size).to(device)
    if model_name == "Mamba":
        net = nn.Sequential(pureMamba(seq_len=1, d_model=in_features, median_size=median_size, state_size=state_size,
                                      output_size=output_size)).to(device)
    if model_name == "Transformer":
        net = nn.Sequential(Transformer(seq_len=1, d_model=in_features, median_size=median_size, state_size=state_size,
                                      output_size=output_size)).to(device)
    if model_name == "MLP":
        net = nn.Sequential(baselinemlp(in_features=in_features, median_size=median_size, output_size = output_size)).to(device)
    if model_name == "TransMamba":
        net = nn.Sequential(TransMamba(seq_len=1, d_model=in_features, median_size=median_size, state_size=state_size,
                                      output_size=output_size)).to(device)
    if model_name == "Jamba":
        net = nn.Sequential(Jamba(seq_len=1, d_model=in_features, median_size=median_size, state_size=state_size,
                                      output_size=output_size)).to(device)
    if model_name == "DenseNet":
        net = nn.Sequential(DenseNet(input_features = in_features,layer_num=(1,1,1,1),growth_rate=1,init_features=median_size,in_channels=median_size,classes=3)).to(device)
    if model_name == "ConvMamba":
        net = ConvMamba(seq_len=1, d_model=in_features, median_size=median_size, state_size=state_size,
                        output_size=output_size).to(device)
    if model_name == "CHD_Paper_Method":
        net = CHD_Paper_Method(seq_len=1, d_model=in_features, median_size=median_size, state_size=state_size,
                        output_size=output_size).to(device)
    if model_name == "CHD_2024_Paper":
        net = CHD_2024_Paper(seq_len=1, d_model=in_features, median_size=median_size, state_size=state_size,
                        output_size=output_size).to(device)

    return net

def write_to_excel(file_path, data):
    try:
        # 打开 Excel 文件
        data = data.item()
        workbook = openpyxl.load_workbook(file_path)
        # 选择第一个工作表
        sheet = workbook.active
        # 获取已有数据的最大行数
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1).value = data
        # 保存修改后的 Excel 文件
        workbook.save(file_path)
    except Exception as e:
        print(f"出现错误: {e}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--bs', type=int, default=4)
    parser.add_argument('--os', type=int, default=2)
    parser.add_argument('--ss', type=int, default=512)
    parser.add_argument('--ms', type=int, default=256)
    parser.add_argument('--model', type=str, default='E_D_Mamba')
    parser.add_argument('--dataset', type=str, default='ICC_rms')
    parser.add_argument('--K', type=int, default=1)
    parser.add_argument('--parameters_file', type=str, default='34.pth')
    parser.add_argument('--feature_analysis', type=bool, default=True)
    parser.add_argument('--seed', type=int, default=0)
    parser.add_argument('--feature_id', type=int, default=2)
    args = parser.parse_args()

    model_name = args.model
    K = args.K
    dataset_name = args.dataset
    state_size = args.ss
    median_size = args.ms
    output_size = args.os
    feature_id = args.feature_id
    model_path = f'E:\\train_result\\{dataset_name}/{model_name}/{args.parameters_file}'
    feature_analysis = args.feature_analysis
    seed = args.seed
    seed_everything(seed)

    mat, labels = read_mouse(dataset_name)
    i = 0

    print(labels)
    binary_labels = np.where(labels == 'Astrocytes', 0, 1)
    print(binary_labels)

    unique_chars, integer_labels = np.unique(binary_labels, return_inverse=True)

    X_train, X_test, y_train, y_test = train_test_split(mat, integer_labels, test_size=0.2, random_state=19)
    print(X_train.shape)
    print(X_test.shape)



    train_label = torch.tensor(
        y_train
    ).to(device)

    test_data = torch.tensor(
        X_test, dtype=torch.float32
    ).to(device)

    test_label = torch.tensor(
        y_test
    ).to(device)
    in_features = X_train.shape[1]
    # net = get_net(in_features, state_size, median_size, 2, model_name)

    # net.load_state_dict(torch.load(model_path))
    net = torch.load(model_path)
    net.eval()
    with torch.no_grad():
        print(feature_id)
        if feature_id < in_features:
            np.random.shuffle(X_train[:, feature_id])
            train_data = torch.tensor(
                X_train, dtype=torch.float32
            ).to(device)
            pred_c, _, x_triplet, x_encoder, x_decoder = net(train_data)
        else:
            train_data = torch.tensor(
                X_train, dtype=torch.float32
            ).to(device)
            pred_c, _, x_triplet, x_encoder, x_decoder = net(train_data)
        loss = nn.CrossEntropyLoss()
        loss_for_analysis = loss(pred_c, train_label)

    write_to_excel(f'../data/SCCML_data/analysis.xlsx', loss_for_analysis)